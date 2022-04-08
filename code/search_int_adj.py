import os

import numpy as np
from jmetal.algorithm.multiobjective import RandomSearch
from jmetal.algorithm.singleobjective import GeneticAlgorithm
from jmetal.operator import IntegerPolynomialMutation
from jmetal.operator import IntegerSBXCrossover
from jmetal.operator.selection import BestSolutionSelection
from jmetal.util.termination_criterion import StoppingByEvaluations
from jmetal.algorithm.multiobjective.nsgaii import NSGAII
from openpyxl import Workbook
from INTAdjProblem import TestingProblem
from jmetal.util.solution import print_function_values_to_file, print_variables_to_file
from tool import print_gen_test_count_to_file, print_variables_to_excel_sheet_int, print_function_values_to_excel_sheet, \
    get_mutant_difficult_list


def search_int(program_name, difficult_level, input_bit, output_bit, mutant_num, k, file_name, alg):
    # input_bit = 10 
    # output_bit = 10 
    # mutant_num = 10
    # k = 102
    # difficult_level = 'easy'
    wb = Workbook()
    test_sheet1 = wb.create_sheet('gen_test', 0)
    test_sheet2 = wb.create_sheet('gen_test_count', 1)
    test_sheet3 = wb.create_sheet('FUN_test', 2)
    test_sheet4 = wb.create_sheet('VAR_test', 3)
    test_sheet5 = wb.create_sheet('objective log', 4)

    mutant_difficult_list = get_mutant_difficult_list(program_name, difficult_level)

    for i in range(mutant_num):
        test_sheet1.append([0] * pow(2, input_bit)) 

    problem = TestingProblem(k, program_name, difficult_level, input_bit, output_bit, mutant_num, test_sheet1,
                             test_sheet5)

    max_evaluations = 20000

    if alg == 'Search':
        algorithm = NSGAII(
            problem=problem,
            population_size=10,  # 10
            offspring_population_size=10,
            mutation=IntegerPolynomialMutation(1.0 / problem.number_of_variables, distribution_index=20),
            crossover=IntegerSBXCrossover(probability=0.9, distribution_index=20),  # 0.9, 20
            selection=BestSolutionSelection(),
            termination_criterion=StoppingByEvaluations(max_evaluations=max_evaluations)  # 1000
        )
    elif alg == "RS":
        algorithm = RandomSearch(
            problem=problem,
            termination_criterion=StoppingByEvaluations(max_evaluations=max_evaluations)
        )
    else:
        print("Search algorithm " + alg + " not supported!")
        return


    algorithm.run()

    print_gen_test_count_to_file(test_sheet1, test_sheet2, input_bit, mutant_difficult_list)

    front = algorithm.get_result()
    # Save results to file
    print_function_values_to_file(front, 'NSGAII outcome/FUN' + file_name + algorithm.label)
    print_variables_to_file(front, 'NSGAII outcome/VAR' + file_name + algorithm.label)

    print_function_values_to_excel_sheet(front, test_sheet3)
    print_variables_to_excel_sheet_int(front, test_sheet4)

    folderResults = 'mutantNotKilledMatrix'
    # If folder doesn't exist, then create it.
    if not os.path.isdir(folderResults):
        os.makedirs(folderResults)

    # wb.save(folderResults + '/' + + 'bin_test.xlsx')
    wb.save(folderResults + '/' + file_name + '.xlsx')

    print(f'Algorithm: ${algorithm.get_name()}')
    print(f'Problem: ${problem.get_name()}')
    print(f'Computing time: ${algorithm.total_computing_time}')
