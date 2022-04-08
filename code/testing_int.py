import collections
import importlib

import numpy as np
import rpy2.robjects as robjects

from openpyxl import Workbook
from programs import CE, AS, SM, BV, QRAM, IQFT
import csv


def dec2bin(n, dec2bin_param):
    a = 1
    list = []
    while a > 0:
        a, b = divmod(n, 2)
        list.append(str(b))
        n = a
    s = ""
    for i in range(len(list) - 1, -1, -1):
        s += str(list[i])
    s = s.zfill(dec2bin_param)
    return s


def wrong_output(i, right_output):
    set_output = set(right_output)
    if i not in set_output:
        return True  # existing wrong output
    return False


def ReadTxtName(rootdir):
    lines = []
    with open(rootdir, 'r') as file_to_read:
        while True:
            line = file_to_read.readline()
            if not line:
                break
            line = line.strip('\n')
            lines.append(line)
    for i in range(len(lines)):
        lines[i] = int(lines[i])
    return lines


def Unique(input, i):
    counts = collections.Counter(input)
    if counts[input[i]] == 1:
        return True
    else:
        return False

def fitness(input, program_name, difficult_level, program_input_num, program_output_num, mutant_num, test_sheet1, test_sheet5):
    module = importlib.import_module('programs.' + str(program_name))
    not_kill_sum = 0
    for m in range(1,mutant_num+1): #mutant program
        print("M"+str(m))
        flag_kill = False 
        input_kill = False 
        method = program_name + '_' + difficult_level + "_M" + str(m)
        run_method = getattr(module, method)
        for i in range(len(input)):
            count_times = 0
            right_output = []

            right_output_str = []

            p = []
            flag_wrong = False 
            if input[i] < pow(2, program_input_num):
                judge = test_sheet1.cell(row=m,column=input[i]+1).value
                if judge == 0:
                    pt_method = getattr(module, "probabilityComputing") 
                    pt = pt_method(input[i])
                    for k in range(len(pt)):
                        if pt[k] > 1e-4:
                            count_times += 1
                            right_output.append(k)
                            right_output_str.append(dec2bin(k,program_output_num))
                            p.append(pt[k])
                    results = run_method(input[i],count_times)
                    # print(dec2bin(i,program_input_num))
                    # print(results)
                    # print(right_output_str)

                    for k in range(len(pt)):
                        k_s = dec2bin(k, program_output_num)
                        if k_s in results:
                            if wrong_output(k, right_output) == True: 
                                flag_wrong = True
                                break
                    if flag_wrong == False: 
                        if count_times == 1:
                            test_sheet1.cell(row=m,column=input[i]+1).value=-1
                        else:
                            wrong_distribution = 0
                            for re in range(5):
                                fre = []
                                for k in range(len(p)):
                                    k_s = dec2bin(right_output[k], program_output_num)
                                    if k_s in results:
                                        fre.append(results[k_s])
                                    else:
                                        fre.append(0)
                                p = np.array(p)
                                fre = np.array(fre)
                                p = robjects.FloatVector(p)
                                fre = robjects.FloatVector(fre)
                                robjects.r('''
                                       chitest<-function(observed,theoretical){
                                           test_result <- chisq.test(x = observed,p = theoretical)
                                           pvalue = test_result$p.value
                                           return (pvalue)
                                       }
                                ''')
                                t = robjects.r['chitest'](fre,p)
                                if t[0] < 0.01: 
                                    wrong_distribution += 1
                                results = run_method(input[i],count_times)
                            if wrong_distribution >= 3:
                                test_sheet1.cell(row=m, column=input[i] + 1).value = 1
                                input_kill = True
                                break
                            else:
                                test_sheet1.cell(row=m, column=input[i] + 1).value = -1
                    else:
                        # print("wrong output")
                        test_sheet1.cell(row=m, column=input[i] + 1).value = 1
                        input_kill = True
                        break
                elif judge == 1:
                    input_kill = True
                    break

        for j in range(pow(2, program_input_num)):
            if test_sheet1.cell(row=m,column=j+1).value == 1:
                flag_kill = True
                break
        
        if flag_kill == False:
            count_pass = 0
            for j in range(pow(2,program_input_num)):
                if test_sheet1.cell(row=m,column=j+1).value == -1:
                    count_pass += 1
        if input_kill == True:
            not_kill_sum += 0
            print(0)
        else:
            if flag_kill == True:
                not_kill_sum += 1
                print(1)
            else:
                value = 1 - (count_pass/pow(2,program_input_num))
                not_kill_sum += value
                print(value)
    count_suite_size = 0
    for i in range(len(input)):
        if input[i]<pow(2, program_input_num):
            count_suite_size += 1
    log_info = [count_suite_size, not_kill_sum]
    test_sheet5.append(log_info)
    return not_kill_sum
