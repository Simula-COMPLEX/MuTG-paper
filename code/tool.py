from openpyxl import Workbook, load_workbook


def print_function_values_to_excel_sheet(solutions, test_sheet):
    solution_idx = 0
    for solution in solutions:
        function_value_idx = 0
        for function_value in solution.objectives:
            test_sheet.cell(row=solution_idx + 1, column=function_value_idx + 1).value = function_value
            function_value_idx += 1
        solution_idx += 1


def print_variables_to_excel_sheet(solutions, test_sheet):
    solution_idx = 0
    for solution in solutions:
        for variables in solution.variables:
            variable_idx = 0
            for variable in variables:
                test_sheet.cell(row=solution_idx + 1, column=variable_idx + 1).value = str(variable)
                variable_idx += 1
        solution_idx += 1

def print_variables_to_excel_sheet_int(solutions, test_sheet):
    # solution_idx = 0
    # for solution in solutions:
    #     for variables in solution.variables:
    #         variable_idx = 0
    #         for variable in variables:
    #             test_sheet.cell(row=solution_idx + 1, column=variable_idx + 1).value = str(variable)
    #             variable_idx += 1
    #     solution_idx += 1
    solution_idx = 0
    for solution in solutions:
        variable_idx = 0
        for variable in solution.variables:
            test_sheet.cell(row=solution_idx + 1, column=variable_idx + 1).value = str(variable)
            variable_idx += 1
        solution_idx += 1


def print_gen_test_count_to_file(test_sheet1, test_sheet2, input_bit, mutant_difficult_list):
    all_rows = test_sheet1.rows
    all_rows_tuple = tuple(all_rows)
    print("len(all_rows_tuple)")
    print(len(all_rows_tuple))
    test_sheet2.cell(row=1, column=1).value = "mutant"
    test_sheet2.cell(row=1, column=2).value = "-1"
    test_sheet2.cell(row=1, column=3).value = "1"
    test_sheet2.cell(row=1, column=4).value = "0"
    #test_sheet2.cell(row=1, column=5).value = "ideal number of 1"
    #test_sheet2.cell(row=1, column=6).value = "false-positive"
    row_idx = 1
    for row in all_rows_tuple:
        minus_one_num = 0
        one_num = 0
        zero_num = 0
        test_sheet2.cell(row=row_idx + 1, column=1).value = str(row_idx)
        for cell in row:
            if str(cell.value) == '-1':
                minus_one_num += 1
            elif str(cell.value) == '1':
                one_num += 1
            elif str(cell.value) == '0':
                zero_num += 1
        print(minus_one_num)
        print(one_num)
        print(zero_num)
        test_sheet2.cell(row=row_idx + 1, column=2).value = str(minus_one_num)
        test_sheet2.cell(row=row_idx + 1, column=3).value = str(one_num)
        test_sheet2.cell(row=row_idx + 1, column=4).value = str(zero_num)
        # print("row_idx=")
        # print(row_idx)
        # print(mutant_difficult_list[row_idx])
        # print(int(pow(2, input_bit) * mutant_difficult_list[row_idx]))
        #test_sheet2.cell(row=row_idx + 1, column=5).value = str(int(pow(2, input_bit) * mutant_difficult_list[row_idx - 1]))
        #test_sheet2.cell(row=row_idx + 1, column=6).value = str(one_num - int(pow(2, input_bit) * mutant_difficult_list[row_idx - 1]))
        row_idx += 1
